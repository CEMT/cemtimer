__author__ = 'Christopher Taylor'
__email__ = 'chris@cemt.co.uk'
__status__ ='Development'
__version__='0.01'
from PyQt4 import QtCore, QtGui
import sys
from openpyxl import Workbook, load_workbook
import datetime
import ConfigParser
import os
import module_locator
#Timer, update every second, acommodate hh:mm:ss
#Auto Hotkey script that rips ticket number from browser into clipboard, program saves current time, time since last save, and ticket number
#to excel file.

class GetContent(QtCore.QThread):
    """This class pulls content from the weathermap and feeds it into a database created in the appdir."""
    def __init__(self):
        """Initiating the get content thread"""
        QtCore.QThread.__init__(self)

class CEMTimer(QtGui.QMainWindow):
    def __init__(self):
        global appDir
        global confIni
        super(CEMTimer,self).__init__()
        self.addComment = 0
        self.setMovable = False
        parse = ConfigParser.SafeConfigParser()
        parse.read(confIni)
        try:
            self.savedir = parse.get('User','savedir')
            x = int(parse.get('User','xpos'))
            y = int(parse.get('User','ypos'))
            if os.path.exists(self.savedir) is not True:
                os.mkdir(self.savedir)
            self.clock = parse.get('User','timer')
        except:
            self.savedir = parse.get('DEFAULT','savedir')
            x = int(parse.get('DEFAULT','xpos'))
            y = int(parse.get('DEFAULT','ypos'))
        self.msec = 0
        self.sec = 0
        self.min = 0
        self.hour = 0
        self.lmsec = 0
        self.lmin = 0
        self.lsec = 0
        self.lhour = 0
        self.build_excel()
        self.build_frame()
        self.move(x,y)
        self.setWindowIcon(QtGui.QIcon(appDir+'\\clock.ico'))
        self.setWindowTitle('CEMTimer')
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint | QtCore.Qt.WindowStaysOnTopHint)
        self.show()

    def mousePressEvent(self,event):
        super(CEMTimer,self).mousePressEvent(event)
        if event.button() == QtCore.Qt.LeftButton:
            self.clickX = event.x()
            self.clickY = event.y()
            self.leftClick = True

    def mouseReleaseEvent(self,event):
        super(CEMTimer,self).mouseReleaseEvent(event)
        parse = ConfigParser.SafeConfigParser()
        parse.read(confIni)
        a = str(self.geometry()).strip('PyQt4.QtCore.QRect(').strip(')').split(',')
        x = int(a[0])
        y = int(a[1])
        parse = ConfigParser.SafeConfigParser()
        parse.read(confIni)
        parse.set('User','xpos','%s'%x)
        parse.set('User','ypos','%s'%y)
        with open(confIni, 'wb') as configfile:
            parse.write(configfile)
        self.leftClick = False

    def mouseMoveEvent(self,event):
        super(CEMTimer,self).mouseMoveEvent(event)
        x = event.globalX()-self.clickX
        y = event.globalY()-self.clickY
        if self.setMovable:
            if self.leftClick: self.move(x,y)

    def contextMenuEvent(self, event):
        menu = QtGui.QMenu(self)
        moveAction = menu.addAction("Move")
        quitAction = menu.addAction("Quit")
        action = menu.exec_(self.mapToGlobal(event.pos()))
        if action == quitAction:
            QtCore.QCoreApplication.instance().quit()
        if action == moveAction:
            if self.setMovable:
                self.setMovable = False
            else:
                self.setMovable = True


    def build_excel(self):
        self.curdate = str(datetime.datetime.now())[0:10]
        self.wbname = "%s%s_Time_Log.xlsx"%(self.savedir,self.curdate)
        try:
            self.wb = load_workbook(self.wbname)
            self.ws = self.wb.active
            rows = self.ws.rows
            lastrow = rows[::-1]
            lastrow = str(lastrow[0][0]).replace("<Cell Time Log.A","").replace(">","")
            self.row = int(lastrow)+1
            self.ws['A%s'%self.row] = ''
            self.ws['B%s'%self.row] = ''
            self.ws['C%s'%self.row] = ''
            self.ws['D%s'%self.row] = 'Application Restarted'
        except:
            print 'excepted, intentional WB doesnt exist'
            self.row = 1
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Time Log"
            self.ws['A1'] = 'Ticket Ref'
            self.ws['B1'] = 'Time Spent'
            self.ws['C1'] = 'Running Time'
            self.ws['D1'] = 'Time when logged'
            self.ws['E1'] = 'Time Serials'
            self.ws['F1'] = 'Comment'
            self.ws['G1'] = 'Total Time Logged'

    def build_frame(self):
        self.vbox = QtGui.QVBoxLayout()
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.update_time)
        tophbox = QtGui.QHBoxLayout()

        self.totaltimer = QtGui.QRadioButton('View Total')
        self.totaltimer.setChecked(False)
        self.sincetimer = QtGui.QRadioButton('Time Spent')
        self.sincetimer.setChecked(True)
        tophbox.addWidget(self.totaltimer)
        tophbox.addWidget(self.sincetimer)
        self.vbox.addLayout(tophbox)
        self.make_lcd()
        self.make_buttons()
        root = QtGui.QWidget()
        root.setLayout(self.vbox)
        self.setCentralWidget(root)

    def make_lcd(self):
        self.lcd = QtGui.QLCDNumber()
        self.lcd.setDigitCount(11)
        self.timer.start(0)
        self.vbox.addWidget(self.lcd)

    def make_buttons(self):
        hbox = QtGui.QHBoxLayout()
        logcomment = QtGui.QPushButton("Comment")
        logcomment.setShortcut('Ctrl+9')
        logcomment.clicked.connect(self.log_comment)

        logtime = QtGui.QPushButton("Log")
        logtime.setAutoDefault(True)
        logtime.setShortcut('Ctrl+0')
        logtime.clicked.connect(self.log_time)

        hbox.addWidget(logtime)
        hbox.addWidget(logcomment)
        self.vbox.addLayout(hbox)

    def log_comment(self):
        self.addComment = 1
        self.comment = 'words'
        popup = QtGui.QDialog()
        popup.setModal(True)
        text, ok = QtGui.QInputDialog.getText(self,'popup','Enter Comment:')
        if ok:
            self.comment = str(text)
            self.log_time()

    def log_time(self):
        self.lmin = 0
        self.lsec = 0
        self.lhour = 0
        cb = QtGui.QApplication.clipboard()
        text = cb.text()
        try:
            text = int(text)
        except:
            text = str(text)
        times = []
        times.append(self.time)
        self.row += 1
        totallogged = self.sum_time('B')
        try:
            self.ws['A%s'%self.row] = text
            self.ws['B%s'%self.row] = str(self.ltime)
            self.ws['C%s'%self.row] = str(self.time)
            self.ws['D%s'%self.row] = str(datetime.datetime.now())[0:19]
            self.ws['E%s'%self.row] = str('=TIMEVALUE(B%s)'%self.row)
            if self.addComment == 1:
                self.ws['F%s'%self.row] = str(self.comment)
                self.addComment = 0
            self.ws['G2'] = str('=SUM(E2:E%s)'%self.row)
            f2cell = self.ws.cell('G2')
            f2cell.number_format = '[h]:mm:ss'
            self.wb.save(self.wbname)
        except IOError:
            pass

    def sum_time(self,column):
        string = '=%s2'%column
        for i in range(2,self.row):
            string = string+'+%s%s'%(column,i)
        return string

    def update_time(self):
        def ttime(self):
            if self.msec < 99:
                self.msec += 1
            else:
                self.msec = 0
                if self.sec < 59:
                    self.sec += 1
                else:
                    if self.min < 59:
                        self.min += 1
                        self.sec = 0
                    else:
                        self.hour += 1
                        self.min = 0
                        self.sec = 0
        def ctime(self):
            if self.lmsec < 99:
                self.lmsec += 1
            else:
                self.lmsec = 0
                if self.lsec < 59:
                    self.lsec += 1
                else:
                    if self.lmin < 59:
                        self.lmin += 1
                        self.lsec = 0
                    else:
                        self.lhour += 1
                        self.lmin = 0
                        self.lsec = 0

        ttime(self)
        ctime(self)

        msec = str(self.msec).zfill(2)
        sec = str(self.sec).zfill(2)
        min = str(self.min).zfill(2)
        hour = str(self.hour).zfill(2)
        lmsec = str(self.lmsec).zfill(2)
        lsec = str(self.lsec).zfill(2)
        lmin = str(self.lmin).zfill(2)
        lhour = str(self.lhour).zfill(2)
        self.time = "{0}:{1}:{2}:{3}".format(hour,min,sec,msec)
        self.ltime = "{0}:{1}:{2}".format(lhour,lmin,lsec)
        if self.totaltimer.isChecked():
            self.lcd.display(self.time)
        elif self.sincetimer.isChecked():
            self.lcd.display(self.ltime+':%s'%lmsec)
        self.timer.start(10)

    def destroyLayout(self, dlayout):
        def deleteItems(layout):
            if layout is not None:
                while layout.count():
                    item = layout.takeAt(0)
                    widget = item.widget()
                    if widget is not None:
                        widget.deleteLater()
                    else:
                        deleteItems(item.layout())
        deleteItems(dlayout)

appDir = str(module_locator.module_path())
if os.path.exists(appDir) is not True:
    os.mkdir(appDir)

confDir = os.path.expanduser("~\\cemtimer")
if os.path.exists(confDir) is not True:
    os.mkdir(confDir)

confIni = confDir+'\\config.ini'
print confIni
userdir = os.path.expanduser("~\\")
config = ConfigParser.SafeConfigParser(defaults={'original_version':'%s'%__version__,
                                                 'savedir':'%s'%confDir,
                                                 'xpos':'500',
                                                 'ypos':'500',
                                                 'timer':'total'})


try:
    config.read(confIni)
except:
    pass
for section in ['User']:
    if config.has_section(section) is True:
        print('config section already created, skipping')
    else:
        print 'adding section'
        config.add_section('User')

with open(confIni,'wb') as configfile:
    config.write(configfile)

def main():
    global appDir
    global confIni
    app = QtGui.QApplication(sys.argv)
    cemtimer = CEMTimer()
    cemtimer.show()
    app.exec_()

if __name__ == '__main__':
    main()